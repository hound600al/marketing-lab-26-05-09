"""
원본 효율정리 xlsx의 '커버링B2C요약' 시트에서 보정 계수 (AD8, AE5, AD5) 추출.

CLI:
    python extract_correction.py --src <원본 효율정리 xlsx> [--out <json>] [--sheet "커버링B2C요약"]
    또는 --cells AD8,AE5,AD5 로 셀 위치 커스터마이즈.
"""
import argparse
import json
import os
from pathlib import Path

import openpyxl


def main() -> None:
    p = argparse.ArgumentParser(description="원본 효율정리 xlsx의 보정 계수 추출")
    p.add_argument("--src", required=True, help="원본 효율정리 xlsx 경로")
    p.add_argument("--out",
                   default=os.environ.get("TIKTOK_REPORT_DATA_DIR",
                                          str(Path.home() / ".tiktok-airbridge-report" / "data"))
                           + "/_correction_factors.json")
    p.add_argument("--sheet", default="커버링B2C요약")
    p.add_argument("--cells", default="AD8,AE5,AD5",
                   help="추출할 셀 3개 (쉼표 구분, 기본: AD8,AE5,AD5 — 원본 효율정리 xlsx의 X열 함수 참조)")
    args = p.parse_args()

    src = Path(args.src)
    if not src.exists():
        print(f"!!! 파일 없음: {src}")
        return

    cell_names = [c.strip() for c in args.cells.split(",")]
    if len(cell_names) != 3:
        print(f"!!! --cells 는 3개 (현재: {len(cell_names)})")
        return

    wb = openpyxl.load_workbook(src, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"!!! 시트 없음: {args.sheet}, 가용: {wb.sheetnames}")
        return
    ws = wb[args.sheet]

    # JSON 키는 ver3 호환을 위해 AD8/AE5/AD5 이름 사용
    keys = ("AD8", "AE5", "AD5")
    result = {k: ws[name].value for k, name in zip(keys, cell_names)}

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[Extract correction]")
    for k, v in result.items():
        print(f"  {k}: {v}")
    print(f"  → {out}")


if __name__ == "__main__":
    main()
