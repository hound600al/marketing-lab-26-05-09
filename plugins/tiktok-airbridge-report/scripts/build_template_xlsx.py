"""
TikTok 광고 효율 template xlsx 빌드 (함수 기반, 26 컬럼).

CLI:
    python build_template_xlsx.py --start YYYY-MM-DD --end YYYY-MM-DD [옵션...]

옵션:
    --data-dir   JSONL 입력 디렉토리 (기본 env TIKTOK_REPORT_DATA_DIR
                 또는 ~/.tiktok-airbridge-report/data)
    --out-dir    출력 디렉토리 (기본 env TIKTOK_REPORT_OUT_DIR
                 또는 ~/Downloads/{YYYY-MM(end기준)}/tiktok-reports)
    --file-prefix  출력 파일명 접두어 (기본 env TIKTOK_REPORT_FILE_PREFIX
                 또는 'TikTok_효율리포트')

입력 JSONL (data-dir 안):
    _airbridge_data.jsonl, _tiktok_data.jsonl,
    _tiktok_adgroups_meta.jsonl, _tiktok_adgroups_insights.jsonl,
    _ad_friendly_names.jsonl (선택),
    _adgroup_to_memo.json (선택, 이전 주 진행액션 메모),
    _correction_factors.json (선택, 보정 계수)

산출: {out-dir}/{file-prefix}_{YY.MM.DD}-{YY.MM.DD}.xlsx
시트: TikTok 효율 / 설정 / TikTok RAW / Airbridge RAW
"""
from __future__ import annotations

import argparse
import json
import os
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# ----- 팔레트 (HeyDealer 코발트 + 잉크) -----
COBALT = "1F3D8A"
COBALT_LIGHT = "4A6BBE"
COBALT_TINT = "EAF0FB"
INK = "0A0A0A"
MUTE = "6B7280"
WHITE = "FFFFFF"
GROUP_FILL = "4A6BBE"
TOTAL_FILL = "DCE6F7"
MEMO_FILL = "FFF8E1"
BUDGET_INPUT_FILL = "E8F5E9"
SETTING_INPUT_FILL = "FFF3E0"

HEADER_FONT = Font(name="Pretendard", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=COBALT)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Pretendard", size=9, color=INK)
ID_FONT = Font(name="Consolas", size=8, color=MUTE)
GROUP_FONT = Font(name="Pretendard", size=11, bold=True, color=WHITE)
MEMO_FONT = Font(name="Pretendard", size=9, italic=True, color=INK)
TOTAL_FONT = Font(name="Pretendard", size=10, bold=True, color=COBALT)
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_INT = '#,##0;-#,##0;"-"'
FMT_WON = '₩#,##0;-₩#,##0;"-"'
FMT_PCT = '0.00%;-0.00%;"-"'

TIKTOK_RAW_COL = {
    "ad_id": 1, "campaign_name": 2, "adgroup_name": 3, "ad_name": 4,
    "ad_id_dup": 5, "spend": 6, "impressions": 7, "clicks": 8,
    "app_install": 9, "registration": 10, "complete_payment": 11, "currency": 12,
}
AIRBRIDGE_RAW_COL = {
    "ad_creative_id": 1, "campaign": 2, "ad_group": 3, "ad_creative": 4,
    "cost": 5, "installs": 6, "install_users": 7, "signup": 8, "purchase": 9,
}


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="TikTok 광고 효율 template xlsx 빌드")
    p.add_argument("--start", required=True, help="시작일 YYYY-MM-DD")
    p.add_argument("--end", required=True, help="종료일 YYYY-MM-DD")
    p.add_argument("--data-dir",
                   default=os.environ.get("TIKTOK_REPORT_DATA_DIR",
                                          str(Path.home() / ".tiktok-airbridge-report" / "data")),
                   help="JSONL 입력 디렉토리")
    p.add_argument("--out-dir",
                   default=os.environ.get("TIKTOK_REPORT_OUT_DIR"),
                   help="출력 디렉토리 (기본: ~/Downloads/{YYYY-MM(end)}/tiktok-reports)")
    p.add_argument("--file-prefix",
                   default=os.environ.get("TIKTOK_REPORT_FILE_PREFIX", "TikTok_효율리포트"),
                   help="출력 파일명 접두어")
    return p.parse_args()


# ----- 유틸 -----
def load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        return []
    rows = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def load_all(data_dir: Path) -> dict:
    data = {
        "airbridge_ads": load_jsonl(data_dir / "_airbridge_data.jsonl"),
        "tiktok_ads": load_jsonl(data_dir / "_tiktok_data.jsonl"),
        "adgroups_meta": load_jsonl(data_dir / "_tiktok_adgroups_meta.jsonl"),
        "adgroups_ins": load_jsonl(data_dir / "_tiktok_adgroups_insights.jsonl"),
    }
    friendly_rows = load_jsonl(data_dir / "_ad_friendly_names.jsonl")
    data["friendly_names"] = {str(r["ad_id"]): (r.get("friendly_name") or r.get("ad_name") or "")
                              for r in friendly_rows}
    mpath = data_dir / "_adgroup_to_memo.json"
    data["adgroup_to_memo"] = json.loads(mpath.read_text(encoding="utf-8")) if mpath.exists() else {}
    cpath = data_dir / "_correction_factors.json"
    data["correction"] = (json.loads(cpath.read_text(encoding="utf-8"))
                          if cpath.exists() else {"AD8": 1.0, "AE5": 1.0, "AD5": 1.0})
    print(f"[Load] tiktok_ads={len(data['tiktok_ads'])} airbridge_ads={len(data['airbridge_ads'])} "
          f"adgroups_meta={len(data['adgroups_meta'])} adgroups_ins={len(data['adgroups_ins'])} "
          f"friendly_names={len(data['friendly_names'])} memos={len(data['adgroup_to_memo'])}")
    return data


def style_header_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ----- 시트: 설정 -----
def build_settings_sheet(wb, correction: dict) -> None:
    ws = wb.create_sheet("설정")
    ws["A1"] = "TikTok 효율 보정 계수 (보정 구매 CPA 계산에 사용)"
    ws["A1"].font = Font(name="Pretendard", size=12, bold=True, color=COBALT)

    rows = [
        ("계수명", "값", "비고", None),
        ("a (커버링B2C요약!AD8)", correction.get("AD8", 1.0),
         "TikTok 효율 시트 X열에서 곱하는 보정 계수", FMT_PCT),
        ("b (커버링B2C요약!AE5)", correction.get("AE5", 1.0), "분모 (좌측)", FMT_PCT),
        ("c (커버링B2C요약!AD5)", correction.get("AD5", 1.0), "분모 (우측)", FMT_PCT),
        (None, None, None, None),
        ("보정 구매 CPA 공식", "= W × B$2 / (B$3 / B$4)",
         "원본 효율정리 xlsx의 X열 함수 그대로", None),
    ]
    for ri, (a, b, c, fmt) in enumerate(rows, start=2):
        ws.cell(row=ri, column=1, value=a)
        ws.cell(row=ri, column=2, value=b)
        ws.cell(row=ri, column=3, value=c)
        if fmt and isinstance(b, (int, float)):
            ws.cell(row=ri, column=2).number_format = fmt
        if a == "계수명":
            for c_idx in (1, 2, 3):
                ws.cell(row=ri, column=c_idx).font = Font(name="Pretendard", size=10,
                                                          bold=True, color=WHITE)
                ws.cell(row=ri, column=c_idx).fill = HEADER_FILL
                ws.cell(row=ri, column=c_idx).alignment = Alignment(horizontal="center",
                                                                     vertical="center")
        elif isinstance(b, (int, float)):
            ws.cell(row=ri, column=2).fill = PatternFill("solid", fgColor=SETTING_INPUT_FILL)
            ws.cell(row=ri, column=2).font = Font(name="Pretendard", size=10, bold=True, color=INK)
            ws.cell(row=ri, column=2).border = BORDER
            ws.cell(row=ri, column=1).border = BORDER
            ws.cell(row=ri, column=3).border = BORDER

    ws["A9"] = "사용 안내"
    ws["A9"].font = Font(name="Pretendard", size=11, bold=True, color=COBALT)
    notes = [
        "• 매주 원본 효율정리 xlsx의 '커버링B2C요약' 시트 AD8 / AE5 / AD5 셀 값을 위 노란 셀에 갱신.",
        "• 두 RAW 시트 데이터를 새 주차로 갱신하면 'TikTok 효율' 시트는 자동 재계산.",
        "• 보정 구매 CPA 컬럼만 이 보정 계수를 사용. 나머지 지표는 RAW 시트 값에서 직접 계산.",
    ]
    for i, n in enumerate(notes, start=10):
        ws.cell(row=i, column=1, value=n)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        ws.cell(row=i, column=1).font = Font(name="Pretendard", size=9, color=INK)

    set_col_widths(ws, {"A": 28, "B": 16, "C": 48, "D": 4})
    ws.sheet_view.showGridLines = False


# ----- 시트: TikTok RAW -----
def build_raw_tiktok(wb, tiktok_ads: list, friendly_names: dict,
                     start_date: date, end_date: date) -> None:
    ws = wb.create_sheet("TikTok RAW")
    ws["A1"] = "기간"
    ws["B1"] = start_date
    ws["B1"].number_format = "yyyy-mm-dd"
    ws["C1"] = "~"
    ws["D1"] = end_date
    ws["D1"].number_format = "yyyy-mm-dd"
    for c_idx in (1, 2, 3, 4):
        ws.cell(row=1, column=c_idx).font = Font(name="Pretendard", size=11, bold=True, color=COBALT)
    ws["F1"] = (f"TikTok 광고매체 raw ({len(tiktok_ads)} rows). "
                "매주 ads.tiktok.com에서 export 후 row 3 이하 데이터만 교체.")
    ws["F1"].font = Font(name="Pretendard", size=9, italic=True, color=MUTE)
    ws.row_dimensions[1].height = 24

    headers = ["광고 ID", "캠페인 이름", "광고 그룹 이름", "광고 이름", "광고 ID(중복)",
               "비용", "노출수", "클릭수(목적지)",
               "앱 설치 수", "등록(앱)", "구매(앱)", "통화"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, len(headers))

    sorted_ads = sorted(tiktok_ads, key=lambda r: -int(r.get("spend", 0) or 0))
    for i, r in enumerate(sorted_ads, start=3):
        ad_id = int(r["ad_id"])
        friendly = friendly_names.get(str(r["ad_id"])) or r.get("ad_name", "")
        ws.cell(row=i, column=1, value=ad_id)
        ws.cell(row=i, column=2, value=r.get("campaign_name", ""))
        ws.cell(row=i, column=3, value=r.get("adgroup_name", ""))
        ws.cell(row=i, column=4, value=friendly)
        ws.cell(row=i, column=5, value=ad_id)
        ws.cell(row=i, column=6, value=int(r.get("spend", 0) or 0)).number_format = FMT_WON
        ws.cell(row=i, column=7, value=int(r.get("impressions", 0) or 0)).number_format = FMT_INT
        ws.cell(row=i, column=8, value=int(r.get("clicks", 0) or 0)).number_format = FMT_INT
        ws.cell(row=i, column=9, value=int(r.get("app_install", 0) or 0)).number_format = FMT_INT
        ws.cell(row=i, column=10, value=int(r.get("registration", 0) or 0)).number_format = FMT_INT
        ws.cell(row=i, column=11, value=int(r.get("complete_payment", 0) or 0)).number_format = FMT_INT
        ws.cell(row=i, column=12, value="KRW").alignment = Alignment(horizontal="center")
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = BORDER
            if ws.cell(row=i, column=c).font.name is None:
                ws.cell(row=i, column=c).font = DATA_FONT

    set_col_widths(ws, {"A": 20, "B": 32, "C": 36, "D": 32, "E": 20,
                        "F": 14, "G": 12, "H": 10, "I": 11, "J": 11, "K": 11, "L": 8})
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


# ----- 시트: Airbridge RAW -----
def build_raw_airbridge(wb, airbridge_ads: list, start_date: date, end_date: date) -> None:
    ws = wb.create_sheet("Airbridge RAW")
    ws["A1"] = "기간"
    ws["B1"] = start_date
    ws["B1"].number_format = "yyyy-mm-dd"
    ws["C1"] = "~"
    ws["D1"] = end_date
    ws["D1"].number_format = "yyyy-mm-dd"
    for c_idx in (1, 2, 3, 4):
        ws.cell(row=1, column=c_idx).font = Font(name="Pretendard", size=11, bold=True, color=COBALT)
    ws["F1"] = (f"Airbridge Actuals raw ({len(airbridge_ads)} rows). Channel=tiktok 필터. "
                "매주 app.airbridge.io에서 export 후 row 3 이하 교체.")
    ws["F1"].font = Font(name="Pretendard", size=9, italic=True, color=MUTE)
    ws.row_dimensions[1].height = 24

    headers = ["Ad Creative ID", "Campaign", "Ad Group", "Ad Creative",
               "Cost (Channel)", "Installs (App)", "Install 유저 수 (App)",
               "회원가입 (App)", "구매 완료 유저 수 (App)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, len(headers))

    sorted_rows = sorted(airbridge_ads, key=lambda r: -int(r.get("cost", 0) or 0))
    for i, r in enumerate(sorted_rows, start=3):
        cid = int(r["ad_creative_id"])
        ws.cell(row=i, column=1, value=cid)
        ws.cell(row=i, column=2, value=r.get("campaign", ""))
        ws.cell(row=i, column=3, value=r.get("ad_group", ""))
        ws.cell(row=i, column=4, value=r.get("ad_creative", ""))
        ws.cell(row=i, column=5, value=int(r.get("cost", 0) or 0)).number_format = FMT_WON
        ws.cell(row=i, column=6, value=int(r.get("installs", 0) or 0)).number_format = FMT_INT
        ws.cell(row=i, column=7, value=int(r.get("install_users", 0) or 0)).number_format = FMT_INT
        ws.cell(row=i, column=8, value=int(r.get("signup", 0) or 0)).number_format = FMT_INT
        ws.cell(row=i, column=9, value=int(r.get("purchase", 0) or 0)).number_format = FMT_INT
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = BORDER
            if ws.cell(row=i, column=c).font.name is None:
                ws.cell(row=i, column=c).font = DATA_FONT

    set_col_widths(ws, {"A": 20, "B": 32, "C": 36, "D": 36,
                        "E": 14, "F": 13, "G": 14, "H": 14, "I": 16})
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


# ----- 시트: TikTok 효율 (메인 view) -----
def build_view_with_formulas(wb, data: dict, ad_rows: list, ag_rows: list,
                              start_date: date, end_date: date) -> None:
    ws = wb.create_sheet("TikTok 효율", 0)

    ws["A1"] = "기간"
    ws["B1"] = start_date
    ws["B1"].number_format = "yyyy-mm-dd"
    ws["C1"] = "~"
    ws["D1"] = end_date
    ws["D1"].number_format = "yyyy-mm-dd"
    ws["F1"] = (f"총 광고세트 {len(ag_rows)}개 · 총 광고 {len(ad_rows)}개 · "
                "매주 두 RAW 시트만 갱신하면 모든 지표 자동 재계산")
    ws["F1"].font = Font(name="Pretendard", size=9, italic=True, color=MUTE)
    for c_idx in (1, 2, 3, 4):
        ws.cell(row=1, column=c_idx).font = Font(name="Pretendard", size=11, bold=True, color=COBALT)
    ws.row_dimensions[1].height = 24

    headers = [
        "광고세트 / 광고", "캠페인", "광고 ID",
        "광고비", "노출수", "클릭수",
        "매체 앱설치", "매체 회원가입", "매체 구매",
        "CPM", "CPC", "CTR",
        "앱설치 CVR", "회원가입 CVR", "구매 CVR",
        "CPI", "회원가입 CPA(매체)", "구매 CPA(매체)",
        "MMP 앱설치", "MMP 회원가입", "MMP 구매",
        "회원가입 CPA(MMP)", "구매 CPA(MMP)",
        "보정 구매 CPA", "일예산", "변경될 예산",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, 2, len(headers))

    ads_by_adgroup: dict[str, list] = {}
    for ad in ad_rows:
        ads_by_adgroup.setdefault(ad["adgroup_name"], []).append(ad)
    sorted_adgroups = sorted(ag_rows, key=lambda r: -r["spend"])

    cur = 3
    cpa_data_ranges: list[tuple[int, int]] = []

    def vt(cell: str, col_idx: int) -> str:
        return f"=IFERROR(VLOOKUP({cell},'TikTok RAW'!$A:$L,{col_idx},0),0)"

    def va(cell: str, col_idx: int) -> str:
        return f"=IFERROR(VLOOKUP({cell},'Airbridge RAW'!$A:$I,{col_idx},0),0)"

    def ratio(n: str, d: str) -> str:
        return f"=IFERROR({n}/{d},0)"

    def cpm(d: str, e: str) -> str:
        return f"=IFERROR({d}*1000/{e},0)"

    def correction(w: str) -> str:
        return f"=IFERROR({w}*설정!$B$2/(설정!$B$3/설정!$B$4),0)"

    for ag in sorted_adgroups:
        ads = sorted(ads_by_adgroup.get(ag["adgroup_name"], []), key=lambda r: -r["spend"])
        if not ads:
            continue
        memo = data.get("adgroup_to_memo", {}).get(ag["adgroup_name"], "")
        status_short = ag["status"].replace("ADGROUP_STATUS_", "")
        budget_str = f"₩{int(ag['budget']):,}" if ag.get("budget") else "-"

        title = f"📁  {ag['adgroup_name']}    |    상태: {status_short}    |    일예산: {budget_str}"
        ws.cell(row=cur, column=1, value=title)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(headers))
        g = ws.cell(row=cur, column=1)
        g.font = GROUP_FONT
        g.fill = PatternFill("solid", fgColor=GROUP_FILL)
        g.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[cur].height = 26
        cur += 1

        if memo:
            ws.cell(row=cur, column=1, value=f"  💬 진행액션: {memo}")
            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(headers))
            mc = ws.cell(row=cur, column=1)
            mc.font = MEMO_FONT
            mc.fill = PatternFill("solid", fgColor=MEMO_FILL)
            mc.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=True)
            ws.row_dimensions[cur].height = 22
            cur += 1

        data_start = cur
        for ad in ads:
            id_cell = f"$C{cur}"
            ws.cell(row=cur, column=1, value=vt(id_cell, TIKTOK_RAW_COL["ad_name"]))
            ws.cell(row=cur, column=2, value=vt(id_cell, TIKTOK_RAW_COL["campaign_name"]))
            ws.cell(row=cur, column=3, value=int(ad["ad_id"]))
            ws.cell(row=cur, column=4, value=vt(id_cell, TIKTOK_RAW_COL["spend"]))
            ws.cell(row=cur, column=5, value=vt(id_cell, TIKTOK_RAW_COL["impressions"]))
            ws.cell(row=cur, column=6, value=vt(id_cell, TIKTOK_RAW_COL["clicks"]))
            ws.cell(row=cur, column=7, value=vt(id_cell, TIKTOK_RAW_COL["app_install"]))
            ws.cell(row=cur, column=8, value=vt(id_cell, TIKTOK_RAW_COL["registration"]))
            ws.cell(row=cur, column=9, value=vt(id_cell, TIKTOK_RAW_COL["complete_payment"]))
            ws.cell(row=cur, column=10, value=cpm(f"D{cur}", f"E{cur}"))
            ws.cell(row=cur, column=11, value=ratio(f"D{cur}", f"F{cur}"))
            ws.cell(row=cur, column=12, value=ratio(f"F{cur}", f"E{cur}"))
            ws.cell(row=cur, column=13, value=ratio(f"G{cur}", f"F{cur}"))
            ws.cell(row=cur, column=14, value=ratio(f"H{cur}", f"F{cur}"))
            ws.cell(row=cur, column=15, value=ratio(f"I{cur}", f"F{cur}"))
            ws.cell(row=cur, column=16, value=ratio(f"D{cur}", f"G{cur}"))
            ws.cell(row=cur, column=17, value=ratio(f"D{cur}", f"H{cur}"))
            ws.cell(row=cur, column=18, value=ratio(f"D{cur}", f"I{cur}"))
            ws.cell(row=cur, column=19, value=va(id_cell, AIRBRIDGE_RAW_COL["installs"]))
            ws.cell(row=cur, column=20, value=va(id_cell, AIRBRIDGE_RAW_COL["signup"]))
            ws.cell(row=cur, column=21, value=va(id_cell, AIRBRIDGE_RAW_COL["purchase"]))
            ws.cell(row=cur, column=22, value=ratio(f"D{cur}", f"T{cur}"))
            ws.cell(row=cur, column=23, value=ratio(f"D{cur}", f"U{cur}"))
            ws.cell(row=cur, column=24, value=correction(f"W{cur}"))
            ws.cell(row=cur, column=25, value="")
            ws.cell(row=cur, column=26, value="")

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=cur, column=c)
                cell.border = BORDER
                if c == 3:
                    cell.font = ID_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c == 1:
                    cell.font = DATA_FONT
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
                else:
                    cell.font = DATA_FONT
            ws.cell(row=cur, column=4).number_format = FMT_WON
            for col_idx in (5, 6, 7, 8, 9, 19, 20, 21):
                ws.cell(row=cur, column=col_idx).number_format = FMT_INT
            for col_idx in (10, 11, 16, 17, 18, 22, 23, 24):
                ws.cell(row=cur, column=col_idx).number_format = FMT_WON
            for col_idx in (12, 13, 14, 15):
                ws.cell(row=cur, column=col_idx).number_format = FMT_PCT
            cur += 1
        data_end = cur - 1

        ws.cell(row=cur, column=1, value="  TOTAL").font = TOTAL_FONT
        ws.cell(row=cur, column=2, value="-")
        ws.cell(row=cur, column=3, value="-")
        for col_idx in [4, 5, 6, 7, 8, 9, 19, 20, 21]:
            col_l = get_column_letter(col_idx)
            ws.cell(row=cur, column=col_idx,
                    value=f"=SUM({col_l}{data_start}:{col_l}{data_end})")
        ws.cell(row=cur, column=10, value=cpm(f"D{cur}", f"E{cur}"))
        ws.cell(row=cur, column=11, value=ratio(f"D{cur}", f"F{cur}"))
        ws.cell(row=cur, column=12, value=ratio(f"F{cur}", f"E{cur}"))
        ws.cell(row=cur, column=13, value=ratio(f"G{cur}", f"F{cur}"))
        ws.cell(row=cur, column=14, value=ratio(f"H{cur}", f"F{cur}"))
        ws.cell(row=cur, column=15, value=ratio(f"I{cur}", f"F{cur}"))
        ws.cell(row=cur, column=16, value=ratio(f"D{cur}", f"G{cur}"))
        ws.cell(row=cur, column=17, value=ratio(f"D{cur}", f"H{cur}"))
        ws.cell(row=cur, column=18, value=ratio(f"D{cur}", f"I{cur}"))
        ws.cell(row=cur, column=22, value=ratio(f"D{cur}", f"T{cur}"))
        ws.cell(row=cur, column=23, value=ratio(f"D{cur}", f"U{cur}"))
        ws.cell(row=cur, column=24, value=correction(f"W{cur}"))
        ws.cell(row=cur, column=25, value=ag.get("budget", 0) or 0)
        ws.cell(row=cur, column=26, value="")

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=cur, column=c)
            cell.fill = PatternFill("solid", fgColor=TOTAL_FILL)
            cell.font = TOTAL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal or "right",
                                       vertical="center")
        ws.cell(row=cur, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=cur, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=cur, column=3).alignment = Alignment(horizontal="center", vertical="center")
        z = ws.cell(row=cur, column=26)
        z.fill = PatternFill("solid", fgColor=BUDGET_INPUT_FILL)
        z.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=cur, column=4).number_format = FMT_WON
        for col_idx in (5, 6, 7, 8, 9, 19, 20, 21):
            ws.cell(row=cur, column=col_idx).number_format = FMT_INT
        for col_idx in (10, 11, 16, 17, 18, 22, 23, 24):
            ws.cell(row=cur, column=col_idx).number_format = FMT_WON
        for col_idx in (12, 13, 14, 15):
            ws.cell(row=cur, column=col_idx).number_format = FMT_PCT
        ws.cell(row=cur, column=25).number_format = FMT_WON
        ws.cell(row=cur, column=26).number_format = FMT_WON

        cpa_data_ranges.append((data_start, cur))
        cur += 1

        ws.row_dimensions[cur].height = 10
        cur += 1

    for s, e in cpa_data_ranges:
        if s <= e:
            rng = f"W{s}:W{e}"
            rule = ColorScaleRule(start_type="percentile", start_value=10, start_color="63BE7B",
                                  mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                  end_type="percentile", end_value=90, end_color="F8696B")
            ws.conditional_formatting.add(rng, rule)

    set_col_widths(ws, {
        "A": 44, "B": 28, "C": 20,
        "D": 14, "E": 12, "F": 10, "G": 11, "H": 13, "I": 10,
        "J": 11, "K": 11, "L": 9,
        "M": 11, "N": 13, "O": 11,
        "P": 11, "Q": 14, "R": 14,
        "S": 12, "T": 13, "U": 11,
        "V": 14, "W": 14, "X": 14,
        "Y": 13, "Z": 14,
    })
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D3"


def build_ad_joined(data: dict) -> list[dict]:
    mmp_by_id = {str(r["ad_creative_id"]): r for r in data["airbridge_ads"]}
    joined = []
    for t in data["tiktok_ads"]:
        ad_id = str(t["ad_id"])
        m = mmp_by_id.get(ad_id)
        joined.append({
            "ad_id": ad_id, "ad_name": t.get("ad_name", ""),
            "adgroup_name": t.get("adgroup_name", ""),
            "campaign_name": t.get("campaign_name", ""),
            "spend": int(t.get("spend", 0) or 0),
            "impressions": int(t.get("impressions", 0) or 0),
            "clicks": int(t.get("clicks", 0) or 0),
            "ch_app_install": int(t.get("app_install", 0) or 0),
            "ch_registration": int(t.get("registration", 0) or 0),
            "ch_complete_payment": int(t.get("complete_payment", 0) or 0),
            "mmp_installs": int(m.get("installs", 0) or 0) if m else 0,
            "mmp_signup": int(m.get("signup", 0) or 0) if m else 0,
            "mmp_purchase": int(m.get("purchase", 0) or 0) if m else 0,
        })
    seen = {r["ad_id"] for r in joined}
    for m in data["airbridge_ads"]:
        cid = str(m["ad_creative_id"])
        if cid not in seen:
            joined.append({
                "ad_id": cid, "ad_name": m.get("ad_creative", ""),
                "adgroup_name": m.get("ad_group", ""),
                "campaign_name": m.get("campaign", ""),
                "spend": int(m.get("cost", 0) or 0),
                "impressions": 0, "clicks": 0,
                "ch_app_install": 0, "ch_registration": 0, "ch_complete_payment": 0,
                "mmp_installs": int(m.get("installs", 0) or 0),
                "mmp_signup": int(m.get("signup", 0) or 0),
                "mmp_purchase": int(m.get("purchase", 0) or 0),
            })
    joined.sort(key=lambda r: -r["spend"])
    return joined


def build_adgroup_joined(data: dict) -> list[dict]:
    meta_by_id = {str(r["adgroup_id"]): r for r in data["adgroups_meta"]}
    mmp_by_name: dict[str, dict] = {}
    for m in data["airbridge_ads"]:
        key = (m.get("ad_group", "") or "").strip()
        if not key:
            continue
        agg = mmp_by_name.setdefault(key, {"installs": 0, "signup": 0, "purchase": 0})
        agg["installs"] += int(m.get("installs", 0) or 0)
        agg["signup"] += int(m.get("signup", 0) or 0)
        agg["purchase"] += int(m.get("purchase", 0) or 0)
    rows = []
    for ins in data["adgroups_ins"]:
        aid = str(ins["adgroup_id"])
        meta = meta_by_id.get(aid, {})
        name = meta.get("adgroup_name", "")
        mmp = mmp_by_name.get(name, {"installs": 0, "signup": 0, "purchase": 0})
        rows.append({
            "adgroup_id": aid, "adgroup_name": name,
            "campaign_id": meta.get("campaign_id", ""), "status": meta.get("status", ""),
            "budget_mode": meta.get("budget_mode", ""), "budget": float(meta.get("budget", 0) or 0),
            "optimization_event": meta.get("optimization_event", ""),
            "spend": int(ins.get("spend", 0) or 0),
            "impressions": int(ins.get("impressions", 0) or 0),
            "clicks": int(ins.get("clicks", 0) or 0),
            "mmp_installs": mmp["installs"], "mmp_signup": mmp["signup"],
            "mmp_purchase": mmp["purchase"],
        })
    rows.sort(key=lambda r: -r["spend"])
    return rows


def main() -> None:
    args = _parse_args()
    start_date = datetime.strptime(args.start, "%Y-%m-%d").date()
    end_date = datetime.strptime(args.end, "%Y-%m-%d").date()
    data_dir = Path(args.data_dir)
    out_dir = Path(args.out_dir) if args.out_dir else \
        Path.home() / "Downloads" / end_date.strftime("%Y-%m") / "tiktok-reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / (f"{args.file_prefix}_{start_date.strftime('%y.%m.%d')}-"
                     f"{end_date.strftime('%y.%m.%d')}.xlsx")

    data = load_all(data_dir)
    ad_rows = build_ad_joined(data)
    ag_rows = build_adgroup_joined(data)
    print(f"[Join] ad_rows={len(ad_rows)} adgroup_rows={len(ag_rows)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_view_with_formulas(wb, data, ad_rows, ag_rows, start_date, end_date)
    build_settings_sheet(wb, data["correction"])
    build_raw_tiktok(wb, data["tiktok_ads"], data["friendly_names"], start_date, end_date)
    build_raw_airbridge(wb, data["airbridge_ads"], start_date, end_date)

    wb.save(out)
    print(f"[Save] {out}")

    total_ad = sum(r["spend"] for r in ad_rows)
    total_ag = sum(r["spend"] for r in ag_rows)
    print(f"[Verify] ad sum={total_ad:,} adgroup sum={total_ag:,}")


if __name__ == "__main__":
    main()
