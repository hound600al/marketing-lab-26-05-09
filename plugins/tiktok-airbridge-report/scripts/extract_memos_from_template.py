"""
이전 주 template xlsx 의 'TikTok 효율' 시트에서 광고세트별 진행액션 메모를 추출.

CLI:
    python extract_memos_from_template.py --src <prev xlsx> [--out <json>] [--sheet "TikTok 효율"]
"""
import argparse
import json
import os
import re
from pathlib import Path

import openpyxl

ADGROUP_PATTERN = re.compile(r"📁\s+(.+?)\s{2,}\|")
MEMO_PATTERN = re.compile(r"💬\s*진행액션[:：]\s*(.+)")


def main() -> None:
    p = argparse.ArgumentParser(description="이전 주 template xlsx의 광고세트별 메모 추출")
    p.add_argument("--src", required=True, help="이전 주 template xlsx 경로")
    p.add_argument("--out",
                   default=os.environ.get("TIKTOK_REPORT_DATA_DIR",
                                          str(Path.home() / ".tiktok-airbridge-report" / "data"))
                           + "/_adgroup_to_memo.json")
    p.add_argument("--sheet", default="TikTok 효율")
    args = p.parse_args()

    src = Path(args.src)
    if not src.exists():
        print(f"!!! 파일 없음: {src}")
        return

    wb = openpyxl.load_workbook(src, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"!!! 시트 없음: {args.sheet}, 가용: {wb.sheetnames}")
        return
    ws = wb[args.sheet]

    memo_by_adgroup: dict[str, str] = {}
    current: str | None = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if not isinstance(a, str):
            current = None
            continue
        m_ag = ADGROUP_PATTERN.search(a)
        if m_ag:
            current = m_ag.group(1).strip()
            continue
        m_memo = MEMO_PATTERN.search(a)
        if m_memo and current:
            memo_by_adgroup[current] = m_memo.group(1).strip()
            current = None

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(memo_by_adgroup, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[Extract] {len(memo_by_adgroup)} adgroup memos → {out}")


if __name__ == "__main__":
    main()
