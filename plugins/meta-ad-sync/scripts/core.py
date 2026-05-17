#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
커버링 효율정리 — 신규 소재 추가 & 서식 적용 통합 스크립트
=============================================================
사용법:
  1. 직접 실행  : config.py 를 채운 뒤 '실행.bat' 더블클릭
  2. Claude 스킬: /add-creatives [ver1_파일_경로]
"""

import sys, re, shutil, os, time, json
sys.stdout.reconfigure(encoding='utf-8')
from copy import copy
import openpyxl

# ============================================================
# ★ 주간 설정 — config.py 에서 덮어쓰거나 직접 수정 ★
# ============================================================

INPUT_FILE  = ""          # 예: "커버링_효율정리_26.05.12(...)ver1.xlsx" 또는 절대 경로
ACTION_NOTE = "신규 추가"  # 진행액션 메모

# 기존 섹션에 소재 삽입
# { "시트명": [(마지막소재행, 합계행, 첫소재행, ["소재명", ...]), ...] }
EXISTING_INSERTS: dict = {}

# 신규 섹션 추가 (시트 맨 아래에 새 그룹 생성)
# { "시트명": [("세트명", ["소재명", ...]), ...] }
NEW_SECTIONS: dict = {}

# ============================================================
# config.py 또는 CLI JSON 인자로 설정 덮어쓰기
# ============================================================
_script_dir = os.path.dirname(os.path.abspath(__file__))
_config_py  = os.path.join(_script_dir, '..', 'config.py')
if os.path.exists(_config_py):
    _ns: dict = {}
    exec(open(_config_py, encoding='utf-8').read(), _ns)
    INPUT_FILE      = _ns.get('INPUT_FILE', INPUT_FILE)
    ACTION_NOTE     = _ns.get('ACTION_NOTE', ACTION_NOTE)
    EXISTING_INSERTS = _ns.get('EXISTING_INSERTS', EXISTING_INSERTS)
    NEW_SECTIONS     = _ns.get('NEW_SECTIONS', NEW_SECTIONS)

if len(sys.argv) > 1 and sys.argv[1].endswith('.json'):
    with open(sys.argv[1], encoding='utf-8') as _f:
        _cfg = json.load(_f)
    INPUT_FILE       = _cfg.get('input_file', INPUT_FILE)
    ACTION_NOTE      = _cfg.get('action_note', ACTION_NOTE)
    EXISTING_INSERTS = _cfg.get('existing_inserts', EXISTING_INSERTS)
    NEW_SECTIONS     = _cfg.get('new_sections', NEW_SECTIONS)
elif len(sys.argv) > 1:
    INPUT_FILE = sys.argv[1]

# ============================================================
# 내부 함수
# ============================================================
HEADERS = [
    None,'세트명','소재별','광고비','노출수','클릭수','앱 설치','회원가입','구매',
    'CPM','CPC','CTR','앱설치 CVR','회원가입 CVR','구매 CVR','CPI',
    '회원가입 CPA','구매 CPA','앱설치','회원가입','구매','회원가입 CPA',
    '구매 CPA','보정 구매 CPA','예산(현재)','예산(변경)',
]

def fix_formula(formula, insert_row, count=1):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    def repl(m):
        col, dol, row_s = m.group(1), m.group(2), m.group(3)
        if dol:
            return m.group(0)
        rn = int(row_s)
        return col + str(rn + count) if rn >= insert_row else m.group(0)
    return re.sub(r'(\$?[A-Z]+)(\$?)(\d+)', repl, formula)

def ad_row_formulas(r):
    return {
        2:  f"=VLOOKUP(C{r},'메타 RAW'!A:G,2,0)",
        4:  f"=VLOOKUP($C{r},'메타 RAW'!$A:$G,3,0)",
        5:  f"=VLOOKUP($C{r},'메타 RAW'!$A:$G,4,0)",
        6:  f"=VLOOKUP($C{r},'메타 RAW'!$A:$G,5,0)",
        7:  f"=VLOOKUP($C{r},'메타 RAW'!$A:$G,6,0)",
        8:  f"=VLOOKUP($C{r},'메타 RAW'!$A:$I,7,0)",
        9:  f"=VLOOKUP($C{r},'메타 RAW'!$A:$I,8,0)",
        10: f'=IFERROR(D{r}/E{r}*1000,0)',
        11: f'=IFERROR(D{r}/F{r},0)',
        12: f'=IFERROR(F{r}/E{r},0)',
        13: f'=IFERROR(G{r}/F{r},0)',
        14: f'=IFERROR(H{r}/F{r},0)',
        15: f'=IFERROR(I{r}/F{r},0)',
        16: f'=IFERROR(D{r}/G{r},0)',
        17: f'=IFERROR(D{r}/H{r},0)',
        18: f'=IFERROR(D{r}/I{r},0)',
        19: f"=IFERROR(VLOOKUP($C{r},'메타 RAW (MMP)'!$B:$G,3,0),0)",
        20: f"=IFERROR(VLOOKUP($C{r},'메타 RAW (MMP)'!$B:$G,5,0),0)",
        21: f"=IFERROR(VLOOKUP($C{r},'메타 RAW (MMP)'!$B:$G,6,0),0)",
        22: f'=IFERROR(D{r}/T{r},0)',
        23: f'=IFERROR(D{r}/U{r},0)',
        24: f'=IFERROR(W{r}*커버링B2C요약!$AD$8/(커버링B2C요약!$AE$5/커버링B2C요약!$AD$5),0)',
        25: f"=VLOOKUP($C{r},'메타 RAW'!$A:$I,9,0)",
    }

def total_row_formulas(t, first, last):
    return {
        3:  'TOTAL',
        4:  f'=SUM(D{first}:D{last})',
        5:  f'=SUM(E{first}:E{last})',
        6:  f'=SUM(F{first}:F{last})',
        7:  f'=SUM(G{first}:G{last})',
        8:  f'=SUM(H{first}:H{last})',
        9:  f'=SUM(I{first}:I{last})',
        10: f'=D{t}*1000/E{t}',
        11: f'=IFERROR(D{t}/F{t},0)',
        12: f'=IFERROR(F{t}/E{t},0)',
        13: f'=IFERROR(G{t}/F{t},0)',
        14: f'=IFERROR(H{t}/F{t},0)',
        15: f'=IFERROR(I{t}/F{t},0)',
        16: f'=IFERROR(D{t}/G{t},0)',
        17: f'=IFERROR(D{t}/H{t},0)',
        18: f'=IFERROR(D{t}/I{t},0)',
        19: f'=SUM(S{first}:S{last})',
        20: f'=SUM(T{first}:T{last})',
        21: f'=SUM(U{first}:U{last})',
        22: f'=IFERROR(D{t}/T{t},0)',
        23: f'=IFERROR(D{t}/U{t},0)',
        24: f'=IFERROR(W{t}*커버링B2C요약!$AD$8/(커버링B2C요약!$AE$5/커버링B2C요약!$AD$5),0)',
        25: f'=Y{first}',
    }

def copy_style(src_cell, dst_cell):
    try:
        dst_cell.font       = copy(src_cell.font)
        dst_cell.fill       = copy(src_cell.fill)
        dst_cell.border     = copy(src_cell.border)
        dst_cell.alignment  = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
    except Exception:
        pass

def apply_row_style(ws, dst_row, src_row, max_col=26):
    """src_row의 서식을 dst_row에 복사"""
    for c in range(1, max_col + 1):
        copy_style(ws.cell(src_row, c), ws.cell(dst_row, c))
    h = ws.row_dimensions[src_row].height
    if h:
        ws.row_dimensions[dst_row].height = h

def find_template_ad_row(ws):
    """시트에서 VLOOKUP이 포함된 첫 번째 소재 행 반환 (서식 템플릿용)"""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 4).value
        if isinstance(v, str) and 'VLOOKUP' in v:
            return r
    return None

def insert_into_existing(ws, last_ad, total_r, first_ad, creatives):
    """기존 섹션에 소재 삽입 (아래에서 위 순서로 호출할 것)"""
    pos = last_ad + 1
    n   = len(creatives)
    ws.insert_rows(pos, n)

    for row in ws.iter_rows(min_row=pos + n):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = fix_formula(cell.value, pos, n)

    new_last = last_ad + n
    for col in ['D','E','F','G','H','I','S','T','U']:
        idx  = openpyxl.utils.column_index_from_string(col)
        cell = ws.cell(total_r, idx)
        if isinstance(cell.value, str) and 'SUM' in cell.value:
            cell.value = re.sub(
                rf'{col}(\d+):?{col}(\d+)',
                lambda m, cl=col, nl=new_last: f'{cl}{m.group(1)}:{cl}{nl}',
                cell.value,
            )

    tmpl = last_ad  # 바로 위 소재 행을 서식 템플릿으로 사용
    for i, creative in enumerate(creatives):
        r = pos + i
        apply_row_style(ws, r, tmpl)
        for col_idx, val in ad_row_formulas(r).items():
            ws.cell(r, col_idx).value = val
        ws.cell(r, 3).value = creative

    return pos

def append_new_section(ws, adset_name, creatives):
    """시트 끝에 신규 섹션 추가 (서식 자동 적용)"""
    tmpl_ad   = find_template_ad_row(ws)
    tmpl_total = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 3).value == 'TOTAL':
            tmpl_total = r
            break

    base = ws.max_row + 2  # 빈 행 1개 띄우기

    # 진행액션 행
    ws.cell(base, 2).value = '진행액션'
    ws.cell(base, 3).value = ACTION_NOTE

    # 헤더 행
    hdr = base + 1
    for i, h in enumerate(HEADERS, 1):
        ws.cell(hdr, i).value = h
    if tmpl_total:
        apply_row_style(ws, hdr, tmpl_total - 1)  # 헤더 위 템플릿 (헤더 행)

    # TOTAL 행
    tot  = hdr + 1
    fa   = tot + 1
    la   = fa + len(creatives) - 1
    for col_idx, val in total_row_formulas(tot, fa, la).items():
        ws.cell(tot, col_idx).value = val
    if tmpl_total:
        apply_row_style(ws, tot, tmpl_total)

    # 소재 행들
    for i, creative in enumerate(creatives):
        r = fa + i
        for col_idx, val in ad_row_formulas(r).items():
            ws.cell(r, col_idx).value = val
        ws.cell(r, 3).value = creative
        if tmpl_ad:
            apply_row_style(ws, r, tmpl_ad)

    return hdr, tot, fa, la

# ============================================================
# 메인 처리
# ============================================================
def resolve_input(path):
    """절대 경로, 상대 경로, 파일명(스크립트 위치 기준) 순서로 시도"""
    if os.path.isabs(path) and os.path.exists(path):
        return path
    cwd_try = os.path.join(os.getcwd(), path)
    if os.path.exists(cwd_try):
        return cwd_try
    script_try = os.path.join(_script_dir, path)
    if os.path.exists(script_try):
        return script_try
    raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")

def main():
    if not INPUT_FILE:
        print("[오류] INPUT_FILE 이 설정되지 않았습니다.")
        print("       config.py 를 채우거나, 인자로 파일 경로를 전달하세요.")
        sys.exit(1)

    src = resolve_input(INPUT_FILE)
    base, ext = os.path.splitext(src)
    dst = base.replace('ver1', 'ver2') if 'ver1' in base else base + '_ver2'
    dst += ext

    if os.path.exists(dst):
        try:
            os.rename(dst, dst)
        except Exception:
            dst = base + f'_신규소재반영_{int(time.time())}' + ext

    shutil.copy2(src, dst)
    print(f'복사 완료: {dst}')

    wb = openpyxl.load_workbook(dst)

    # 1) 기존 섹션 삽입 (아래→위)
    for sheet_name, inserts in EXISTING_INSERTS.items():
        ws = wb[sheet_name]
        for last_ad, total_r, first_ad, creatives in sorted(inserts, key=lambda x: -x[0]):
            insert_into_existing(ws, last_ad, total_r, first_ad, creatives)
            print(f'  [{sheet_name}] {last_ad+1}행에 {len(creatives)}개 삽입 완료')

    # 2) 신규 섹션 추가
    for sheet_name, sections in NEW_SECTIONS.items():
        ws = wb[sheet_name]
        for adset_name, creatives in sections:
            h, t, fa, la = append_new_section(ws, adset_name, creatives)
            print(f'  [{sheet_name}] 신규 섹션 "{adset_name[:30]}" 행{h}~{la} 추가 완료')

    wb.save(dst)
    print(f'\n✅ 완료: {dst}')
    return dst

if __name__ == '__main__':
    main()
